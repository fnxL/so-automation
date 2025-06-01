import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, Alignment


def format_number(ws: Worksheet, startcol=0, endcol=1, format="0"):
    for row in ws.iter_cols(min_col=startcol, max_col=endcol):
        for cell in row:
            cell.number_format = format
    return


def apply_borders(ws: Worksheet, text_center=False):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    text_center = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            if text_center:
                cell.alignment = text_center

    return


def get_df_from_excel(path: str, sheet_name: str | int = 0, **kwargs) -> pd.DataFrame:
    df = pd.read_excel(io=path, engine="calamine", sheet_name=sheet_name, **kwargs)
    df.columns = df.columns.str.strip()
    df.columns = df.columns.str.lower()
    return df
