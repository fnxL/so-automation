import win32com.client as win32
import pandas as pd
import re
from pywinauto.application import Application
from loguru import logger
from ..integrations.sap_connector import start_sap_alert_thread
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, Alignment


class ExcelClient:
    """
    Usage:
        ```python
        try:
            with ExcelClient() as excel:
                excel.open("absolute/path/to/excel/file.xlsx")
                excel.copy_used_range()
        except Exception as e:
            logger.error(f"Error: {e}")
    """

    def __init__(self, logger=logger, visible: bool = True):
        self.logger = logger
        self.visible = visible
        self.excel = None
        self.workbook = None

    def __enter__(self):
        self.logger.info("Connecting to Excel application.")
        try:
            self.excel = win32.GetActiveObject("Excel.Application")
            self.logger.success("Connected to existing Excel instance")
        except Exception:
            self.logger.info("Starting Excel application via COM...")
        try:
            self.excel = win32.Dispatch("Excel.Application")
            self.excel.Visible = self.visible
            self.logger.success("Created new Excel instance")
        except Exception as e:
            self.logger.error(f"Failed to initialize Excel application: {e}")
            self.cleanup()
            raise e

        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.logger.info("Cleaning up Excel application instance...")
        self.cleanup()

    def open(self, path: str):
        if not path:
            raise ValueError("Path is required to open Excel file.")

        self.logger.info(f"Opening workbook: {path}")
        try:
            self.workbook = self.excel.Workbooks.Open(path)
        except Exception as e:
            self.logger.error(f"Failed to open workbook: {e}")
            raise e

        return self

    def run_macro(self, macro_name: str):
        if not self.workbook:
            raise RuntimeError("Workbook is not open. Please open the workbook first")
        sap_alert_thread = start_sap_alert_thread(logger=self.logger)
        try:
            self.logger.info(f"Running macro: {macro_name}")
            self.excel.Application.Run(macro_name)
            self.logger.success(f"Macro '{macro_name}' executed  successfully.")
            sap_alert_thread.join(timeout=5)
            return True
        except Exception as e:
            self.logger.warning(f"An error occurred while running macro: {e}")
            raise e
        finally:
            sap_alert_thread.join(timeout=5)

    def copy_used_range(self, sheet_index=1):
        if not self.workbook:
            raise RuntimeError("Workbook is not open. Please open the workbook first")

        sheet = self.workbook.Sheets(sheet_index)
        sheet.UsedRange.Copy()
        self.logger.info(f"Copied entire used range from Sheet {sheet_index}.")
        return

    def find_and_close_workbooks(
        self,
        title_contains: str,
        save_changes: bool = False,
    ):
        try:
            if self.excel.Workbooks.Count == 0:
                self.logger.info("No workbooks are currently open")
                return False

            dispatch_workbooks = []
            for i in range(1, self.excel.Workbooks.Count + 1):
                wb = self.excel.Workbooks(i)
                if title_contains in wb.Name:
                    dispatch_workbooks.append(wb)
                    self.logger.info(f"Found matching workbook: {wb.Name}")

            # Close found workbooks
            for wb in dispatch_workbooks:
                try:
                    wb_name = wb.Name
                    self.excel.DisplayAlerts = False
                    wb.Close(SaveChanges=save_changes)
                    self.excel.DisplayAlerts = True
                    self.logger.success(f"Successfully closed workbook: {wb_name}")
                except Exception as e:
                    self.logger.error(f"Error closing workbook {wb.Name}: {e}")

            return len(dispatch_workbooks) > 0

        except Exception as e:
            self.logger.error(f"Error while searching for workbooks: {e}")
            return False

    def cleanup(self):
        self.excel.DisplayAlerts = False
        if self.workbook:
            try:
                self.workbook.Close(SaveChanges=True)
                self.logger.info("Workbook closed and saved.")
            except Exception as e:
                self.logger.warning(f"Could not close workbook: {e}")
            self.workbook = None

        if self.excel:
            try:
                self.excel.Quit()
                self.logger.info("Excel application quit.")
            except Exception as e:
                self.logger.warning(f"Could not quit Excel application: {e}")
            self.excel = None


def find_and_close_workbook_fallback(title_contains: str, logger=logger):
    """Uses pywinauto to close the workbook"""
    try:
        app = Application(backend="uia").connect(title_re=f".*{title_contains}.*")
        app.top_window().close()
        logger.info(f"{re.escape(title_contains)}: Workbook closed successfully.")
    except Exception as e:
        logger.error(f"Failed to close workbook/no workbook found: {e}")
        raise e


# openpyxl utilities
def get_df_from_excel(path: str, sheet_name: str | int = 0, **kwargs) -> pd.DataFrame:
    df = pd.read_excel(io=path, engine="calamine", sheet_name=sheet_name, **kwargs)
    df.columns = df.columns.str.strip()
    df.columns = df.columns.str.lower()
    return df


def format_number(ws: Worksheet, startcol=0, endcol=1, format="0"):
    for row in ws.iter_cols(min_col=startcol, max_col=endcol):
        for cell in row:
            cell.number_format = format
    return


def apply_borders(ws: Worksheet, text_center=False):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    text_center = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            if text_center:
                cell.alignment = text_center
    return

def adjust_column_widths(ws: Worksheet, adjustment_factor: float = 1.05):
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        column_letter = column[0].column_letter
        ws.column_dimensions[column_letter].width = (max_length + 2) * adjustment_factor

    return
