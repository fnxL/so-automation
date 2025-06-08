from loguru import logger
from ...integrations.excel import get_df_from_excel, format_number, apply_borders
from ...integrations.pdf_parser import process_pdf
from ...integrations import ExcelClient, OutlookClient
from openpyxl import load_workbook
from datetime import datetime
from dataclasses import dataclass
from typing import List, Tuple
import os
import time


@dataclass
class POData:
    po: int | str
    port_of_shipment: str
    channel_type: str
    sub_channel_type: str
    ship_start_date: datetime
    ship_end_date: datetime
    packing_type: str
    notify: str


class KohlsAutomation:
    def __init__(
        self,
        config,
        source_folder: str,
        logger=logger,
    ):
        self.config = config
        self.logger = logger
        self.source_folder = source_folder
        self.load_config()

        self.macro_wb = None
        self.macro_ws = None
        self.mastersheet_df = None
        self.pis_df = None

    def load_config(self):
        self._macro_path = self.config["macro_path"]
        self._mastersheet_path = self.config["mastersheet_path"]
        self._notify_address = self.config["notify_address"]
        self._design_split = self.config["design_split"]
        self._source_folder_cell = self.config["source_folder_cell"]

    def _create_draft_mails(self, reports):
        with OutlookClient(self.logger) as outlook:
            for report in reports:
                plant = str(report[0])
                report_path = report[1]

                # apply borders first
                wb = load_workbook(filename=report_path)
                ws = wb.active
                apply_borders(ws)
                wb.save(report_path)
                time.sleep(2)

                self.logger.info(f"Copying dispatch report to clipboard: {report_path}")
                with ExcelClient(excel_path=report_path, logger=self.logger) as excel:
                    excel.copy_used_range()

                to = self.config["mail"][plant]["to"]
                cc = self.config["mail"][plant]["cc"]
                subject = self.config["mail"][plant]["subject"]
                body = self.config["mail"][plant]["body_template"]
                self.logger.info(f"Creating email for plant: {plant}")
                outlook.create_draft_mail(
                    to, cc, subject, body, paste_from_clipboard=True
                )

    def _get_pdf_files(self, path: str):
        return [f for f in os.listdir(path) if f.lower().endswith(".pdf")]

    def _parse_ship_date(self, date_str: str):
        return datetime.fromisoformat(date_str)

    def _get_mastersheet_row(self, upc):
        if self.mastersheet_df is None:
            self.mastersheet_df = get_df_from_excel(path=self._mastersheet_path)

        result = self.mastersheet_df.query(f"upc=={upc}")
        if result.empty:
            self.logger.error(f"No mastersheet row found for UPC: {upc}")
            raise ValueError(f"No mastersheet row found for UPC: {upc}")
        return result.iloc[0]

    def _parse_po_metadata(self, po_metadata: dict) -> POData:
        ship_start_date = self._parse_ship_date(po_metadata["ship_start_date"])
        ship_end_date = self._parse_ship_date(po_metadata["ship_end_date"])
        sub_channel_type = (
            "PURE PLAY ECOM" if po_metadata["channel_type"] == "ECOM" else "NIL"
        )
        packing_type = "BULK" if po_metadata["channel_type"] == "RETAIL" else "ECOM"
        notify = (
            self.config["notify_address"]
            if "notify" in po_metadata
            else "2% commission to WUSA"
        )

        return POData(
            po=po_metadata["po"],
            port_of_shipment=po_metadata["port_of_shipment"],
            channel_type=po_metadata["channel_type"],
            ship_start_date=ship_start_date,
            ship_end_date=ship_end_date,
            sub_channel_type=sub_channel_type,
            packing_type=packing_type,
            notify=notify,
        )

    def _prepare_row_data(self, po_data: POData, line_items: List[Tuple]):
        row_data = {}
        for row in line_items:
            _, _, _, qty, _, _, upc = row
            mastersheet_row = self._get_mastersheet_row(upc)
            # Prepare entire excel macro row
            macro_row = self._create_macro_row(
                po_data=po_data,
                qty=qty,
                upc=upc,
                mastersheet_row=mastersheet_row,
            )

            row_group = self._get_row_group_key(mastersheet_row)

            if row_group not in row_data:
                row_data[row_group] = []
            row_data[row_group].append(macro_row)
        return row_data

    def _get_row_group_key(self, mastersheet_row):
        sales_unit = mastersheet_row["sales unit"]
        design = str(mastersheet_row["design"]).lower().strip()

        if design in self.config["design_split"]:
            return f"{design}_{sales_unit}"
        return sales_unit

    def _get_adjusted_po(self, base_po: int | str, design: str):
        if design in self.config["design_split"]:
            return f"{base_po} {design}"
        return base_po

    def _get_pis_data(self, mastersheet_row, po_data: POData):
        if self.pis_df is None:
            self.pis_df = get_df_from_excel(
                path=self._mastersheet_path, sheet_name="PIS"
            )

        sales_unit = mastersheet_row["sales unit"]
        program_name = mastersheet_row["program name"]
        packing_type = po_data.packing_type
        filtered_pis = self.pis_df[
            (self.pis_df["program name"] == program_name)
            & (self.pis_df["sales unit"] == sales_unit)
            & (self.pis_df["packing type"] == packing_type)
        ]

        return {
            "pis": filtered_pis.iloc[0]["pis"] if not filtered_pis.empty else "N/A",
            "f_part": filtered_pis.iloc[0]["f part"]
            if not filtered_pis.empty
            else "N/A",
        }

    def _get_s_part(
        self, plant: int, packing_type: str, sales_unit: str, ship_month: int
    ):
        if plant == 2100:
            if packing_type == "BULK" and sales_unit == "PC":
                return f"ALT{ship_month}"
            elif packing_type == "ECOM" and sales_unit == "PC":
                return f"ALT{ship_month}"
            elif packing_type == "ECOM" and (
                sales_unit == "12 PC SET" or sales_unit == "6 PC SET"
            ):
                return f"ALT{ship_month + 12}"
        return ""

    def _process_single_po(self, pdf_file: str):
        pdf_path = os.path.join(self.source_folder, pdf_file)
        self.logger.info(f"Processing PDF: {pdf_file}")

        po_metadata, line_items = process_pdf(pdf_path=pdf_path)
        po_data = self._parse_po_metadata(po_metadata)
        row_data = self._prepare_row_data(po_data, line_items)
        self.logger.info("Parsed PO File")
        self.logger.info("Writing data to macro worksheet...")
        self._write_rows_to_worksheet(row_data)

    def _write_rows_to_worksheet(self, row_data):
        if self.macro_ws is None:
            self.macro_wb = load_workbook(filename=self._macro_path, keep_vba=True)
            self.macro_ws = self.macro_wb.worksheets[0]

        for group, rows in row_data.items():
            for row in rows:
                self.macro_ws.append(row)
            self.macro_ws.append([""])  # Add a blank row after each group

    def _finalize_macro_file(self):
        format_number(
            self.macro_ws, startcol=32, endcol=32, format="0"
        )  # UPC column long number format

        format_number(
            self.macro_ws, startcol=13, endcol=14, format="DD-MM-YYYY"
        )  # ship dates

        # Set source folder in macro file
        self.macro_ws[self.config["source_folder_cell"]] = self.source_folder

        # Save filled macro
        macro_filename = "FILLED_" + os.path.basename(self._macro_path)
        output_path = os.path.join(self.source_folder, macro_filename)
        self.macro_wb.save(output_path)
        self.logger.info(f'Macro file saved to "{output_path}"')

        return output_path
