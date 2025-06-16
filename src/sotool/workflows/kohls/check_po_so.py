from sotool.integrations.excel import get_df_from_excel
from loguru import logger
from sotool.integrations.excel import apply_borders, format_number, adjust_column_widths
from sotool.workflows.kohls.pdf_parser import parse_po_metadata, get_total_qty_and_value
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
import pandas as pd
import os


def check_po_so(report_path, source_folder, logger=logger):
    report_df = get_df_from_excel(
        path=report_path, sheet_name=0, dtype={"First Pis number": str}
    )
    report_df = report_df.rename(columns={"destination": "packing"})
    if "first pis number" in report_df.columns:
        report_df["first pis number"] = (
            report_df["first pis number"].astype(str).str.lstrip("0")
        )
        report_df["first pis number"] = pd.to_numeric(
            report_df["first pis number"], errors="coerce"
        ).astype("Int64")

    if "cbm" in report_df.columns:
        report_df = report_df.drop(columns=["cbm"])

    report_df["remarks"] = ""

    def split_po(x):
        if isinstance(x, str):
            po = int(x.split()[0])
            return po
        return x

    # temp column for clean po number
    report_df["po_number_clean"] = report_df["buyer po no"].apply(func=split_po)

    # ignore last row for grouping
    grouping_df = report_df.iloc[:-1]
    df = grouping_df.groupby("po_number_clean")

    for po, group in df:
        so_order_qty = round(group["so order qty"].sum(), 2)
        so_value = round(group["so value"].sum(), 2)

        if isinstance(po, str):
            po = po.split(" ")[0]

        po = int(po)
        po_path = os.path.join(source_folder, f"{po}.pdf")

        # Update packing
        po_metadata = parse_po_metadata(po_path)
        report_df.loc[group.index, "packing"] = po_metadata.packing_type
        po_total_value = get_total_qty_and_value(po_path)
        po_total_qty = po_total_value["total_qty"]
        po_total_order_value = po_total_value["order_value"]

        qty_match = int(so_order_qty) == po_total_qty
        value_match = so_value == po_total_order_value

        # Update remarks for all rows in this PO group
        if not qty_match and not value_match:
            report_df.loc[group.index, "remarks"] = "Qty and Value not matching"
        elif not qty_match:
            report_df.loc[group.index, "remarks"] = "Qty not matching"
        elif not value_match:
            report_df.loc[group.index, "remarks"] = "Value not matching"
        else:
            report_df.loc[group.index, "remarks"] = "Qty and Value Match"

        logger.info(f"PO: {po}, Qty Match: {qty_match}, Value Match: {value_match}")

    report_df = report_df.drop(columns=["po_number_clean"])
    report_df.columns = [col.title() for col in report_df.columns]
    file_name = os.path.basename(report_path)
    output_path = os.path.join(source_folder, f"OK_{file_name}")
    report_df.to_excel(output_path, index=False)
    format_excel(output_path)
    return output_path


def format_excel(file_path):
    wb = load_workbook(filename=file_path)
    ws = wb.active
    max_row = ws.max_row
    max_col = ws.max_column

    apply_borders(ws)
    format_number(ws, startcol=10, endcol=11, format="DD-MM-YYYY")

    header_fill = PatternFill(
        start_color="c0c0c0", end_color="c0c0c0", fill_type="solid"
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=max_row, column=col)
        cell.fill = yellow_fill

    # comma separated values
    format_number(ws, startcol=4, endcol=8, format="#,##0.00")
    adjust_column_widths(ws)

    # highlight remarks col
    ErrorFill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for col in ws.iter_cols(min_col=13, max_col=13):
        for cell in col:
            if cell.value is None:
                continue
            if "not" in cell.value.lower():
                cell.fill = ErrorFill

    wb.save(file_path)
